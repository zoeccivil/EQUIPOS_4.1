"""
Generador de reportes PDF para EQUIPOS 4.0
Adaptado para trabajar con Firebase y Firebase Storage
"""

import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.lib import colors
from datetime import datetime
import os
import tempfile
import logging

import logging
from reportlab.platypus import Paragraph, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from xml.sax.saxutils import escape
logger = logging.getLogger(__name__)




# Recomendado al inicio del archivo:
# from datetime import datetime
# try:
#     import pandas as pd  # opcional, solo para compatibilidad
# except ImportError:
#     pd = None

class ReportGenerator:
    """
    Generador de reportes PDF con soporte para conduces desde Firebase Storage.
    """

    def __init__(
        self,
        data=None,
        title: str = "",
        cliente: str = "",
        date_range: str = "",
        currency_symbol: str = "RD$",
        storage_manager=None,
        column_map: dict | None = None,
    ):
        """
        Inicializa el generador de reportes.

        Args:
            data: Lista de diccionarios con los registros (facturas) a incluir en el reporte.
            title: Título del reporte (si está vacío se usa "REPORTE DE ALQUILERES").
            cliente: Nombre del cliente (o "GENERAL" si es estado general).
            date_range: Rango de fechas a mostrar en el encabezado (ej: '2025-01-01 a 2025-11-18').
            currency_symbol: Símbolo de moneda (por defecto 'RD$').
            storage_manager: Instancia de StorageManager (para resolver links de conduces).
            column_map: Mapeo de columnas {clave_dato: 'Etiqueta en PDF'}.
        """
        # Datos base que usa to_pdf
        self.data: list[dict] = list(data or [])
        self.title: str = title or "REPORTE DE ALQUILERES"
        self.cliente: str = cliente or ""
        self.date_range: str = date_range or ""
        self.currency_symbol: str = currency_symbol or "RD$"
        self.storage_manager = storage_manager
        self.column_map: dict = dict(column_map or {})

        # Metadatos
        self.fecha_generacion: str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Estado de cuenta: entradas opcionales que puede setear el llamador (AppGUI)
        # Si no las setean, to_pdf calcula/agrupa lo necesario con fallbacks.
        self.abonos: list[dict] = []                 # lista cruda de abonos (opcional)
        self.abonos_por_fecha: list[tuple] = []      # [(fecha, total_en_fecha)] opcional
        self.total_facturado: float = 0.0
        self.total_abonado: float = 0.0
        self.saldo: float = 0.0

        # Compatibilidad opcional con código legado que esperaba un DataFrame.
        # to_pdf NO depende de self.df; esto es por si lo usas en otros reportes.
        self.df = None
        try:
            if 'pd' in globals() and pd is not None and self.data:
                raw_df = pd.DataFrame([dict(r) for r in self.data])
                if self.column_map and not raw_df.empty:
                    cols_a_usar = [c for c in self.column_map.keys() if c in raw_df.columns]
                    self.df = raw_df[cols_a_usar].rename(columns=self.column_map)
                else:
                    self.df = raw_df
        except Exception:
            # Si pandas no está o falla, seguimos sin self.df
            self.df = None

        # Archivos temporales descargados (si bajas conductos para otros flujos)
        self.temp_files: list[str] = []
    
    def _group_abonos_by_date(self, abonos: list[dict]) -> list[tuple[str, float]]:
        """
        Devuelve [(fecha 'YYYY-MM-DD', total_en_fecha), ...] ordenada por fecha asc.
        """
        acum = {}
        for a in abonos or []:
            fecha = a.get("fecha")
            if not fecha:
                continue
            monto = float(a.get("monto", 0) or 0.0)
            acum[fecha] = acum.get(fecha, 0.0) + monto
        return sorted(acum.items(), key=lambda x: x[0])


    def _resolve_condstorage_url(self, value: str) -> str:
        """
        Convierte 'CondStorage' (path o URL) en una URL final:
        - Si ya es URL, la devuelve.
        - Si es path de Storage y hay storage_manager, genera URL de descarga.
        """
        if not value:
            return ""
        if isinstance(value, str) and value.startswith(("http://", "https://")):
            return value
        try:
            sm = getattr(self, "storage_manager", None)
            if sm:
                return sm.get_download_url(value) or value
            return value
        except Exception:
            return value


    def _make_table_abonos_por_fecha(self, abonos_por_fecha: list[tuple[str, float]], currency_symbol: str = "RD$"):
        """
        Construye una tabla ReportLab: Fecha | Total Abonado
        """
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors

        data = [["Fecha", "Total Abonado"]]
        for fecha, total in abonos_por_fecha or []:
            data.append([fecha, f"{currency_symbol} {total:,.2f}"])

        tbl = Table(data, hAlign="LEFT")
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF1E0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#A35D00")),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#A35D00")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]))
        return tbl


    def _postprocess_row_for_pdf(self, row: dict, column_map: dict) -> list:
        """
        Transforma un dict 'row' en una lista de celdas según column_map.
        - Convierte CondStorage en un link "Ver" si hay URL.
        - Formatea horas y monto.
        """
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import getSampleStyleSheet

        styles = getSampleStyleSheet()
        pdf_row = []
        currency_symbol = getattr(self, "currency_symbol", "RD$")

        for key in column_map.keys():
            val = row.get(key, "")

            if key == "CondStorage":
                url = self._resolve_condstorage_url(val)
                if url and isinstance(url, str) and url.startswith(("http://", "https://")):
                    val = Paragraph(f'<link href="{url}">Ver</link>', styles["BodyText"])
                else:
                    val = ""
            else:
                if key == "horas" and val not in ("", None):
                    try:
                        val = f"{float(val):,.2f}"
                    except Exception:
                        pass
                if key == "monto" and val not in ("", None):
                    try:
                        val = f"{currency_symbol} {float(val):,.2f}"
                    except Exception:
                        pass

            pdf_row.append(val)

        return pdf_row



    def _build_facturas_table(self, column_map, data, font_size=9, page_w=None):
        """
        Construye la tabla de facturas con:
        - Wrapping en columnas flexibles
        - Anchos dinámicos ajustados a la página (self._auto_compute_col_widths)
        - repeatRows=1 para repetir encabezados en cada página
        """
        if not column_map or not data:
            return Paragraph("Sin datos de facturas.", getSampleStyleSheet()["Normal"])

        keys = list(column_map.keys())
        headers = [column_map[k] for k in keys]

        # Filas con wrapping
        rows = self._rows_with_wrapping(column_map, data, font_size=font_size)

        # Calcular anchos dinámicos
        if page_w is None:
            page_w, _ = landscape(LETTER)
        col_widths = self._auto_compute_col_widths(
            column_map,
            data,
            page_w,
            margins=(36, 36),
            font_name="Helvetica",
            font_size=font_size
        )

        tbl = Table([headers] + rows,
                    hAlign="LEFT",
                    colWidths=col_widths,
                    repeatRows=1)  # <--- encabezados se repiten

        # Alinear columnas numéricas
        num_cols = [i for i, k in enumerate(keys) if k in ("horas", "monto")]

        ts = [
            ("FONTSIZE", (0, 0), (-1, -1), font_size),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6F4EA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F7A1F")),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#1F7A1F")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
        for idx in num_cols:
            ts.append(("ALIGN", (idx, 1), (idx, -1), "RIGHT"))

        tbl.setStyle(TableStyle(ts))
        return tbl


    def to_pdf(self, out_path: str):
        """
        Genera el PDF horizontal con:
        - Página(s) de Facturas (tabla con encabezado repetido).
        - PageBreak.
        - Página de Abonos por fecha + Totales.
        - Anexos (conduces) al final.
        """
        try:
            import tempfile, os

            styles = getSampleStyleSheet()
            story = []

            # Datos base
            title = getattr(self, "title", "ESTADO DE CUENTA")
            cliente = getattr(self, "cliente", "")
            date_range = getattr(self, "date_range", "")
            self.currency_symbol = getattr(self, "currency_symbol", "RD$")
            column_map = getattr(self, "column_map", {}) or {}
            data = list(getattr(self, "data", []) or [])

            # Encabezado principal
            story.append(Paragraph(title, styles["Title"]))
            if cliente:
                story.append(Paragraph(f"Cliente: {cliente}", styles["Heading3"]))
            if date_range:
                story.append(Paragraph(f"Periodo: {date_range}", styles["Normal"]))
            story.append(Spacer(1, 10))

            # Si no hay column_map pero hay datos, generar uno básico
            if not column_map and data and isinstance(data[0], dict):
                column_map = {k: k.capitalize() for k in data[0].keys()}

            # Tabla Facturas (puede paginar automáticamente)
            story.append(Paragraph("Facturas", styles["Heading3"]))
            facturas_tbl = self._build_facturas_table(column_map, data, font_size=9)
            story.append(facturas_tbl)

            # Page break para iniciar sección de abonos en página nueva
            story.append(PageBreak())

            # Abonos por fecha (agrupar si no viene agrupado)
            abonos_por_fecha = getattr(self, "abonos_por_fecha", None)
            if not abonos_por_fecha:
                abonos_list = getattr(self, "abonos", [])
                abonos_por_fecha = self._group_abonos_by_date(abonos_list)

            story.append(Paragraph("Abonos por fecha", styles["Heading3"]))
            story.append(self._make_table_abonos_por_fecha(abonos_por_fecha, currency_symbol=self.currency_symbol))
            story.append(Spacer(1, 18))

            # Totales
            total_facturado = float(getattr(self, "total_facturado", 0) or 0)
            total_abonado = float(getattr(self, "total_abonado", 0) or 0)
            saldo = float(getattr(self, "saldo", total_facturado - total_abonado) or 0)

            tot_headers = ["Total Facturas", "Total Abonos", "Saldo"]
            tot_values = [
                f"{self.currency_symbol} {total_facturado:,.2f}",
                f"{self.currency_symbol} {total_abonado:,.2f}",
                f"{self.currency_symbol} {saldo:,.2f}",
            ]
            tot_tbl = Table([tot_headers, tot_values], hAlign="RIGHT", colWidths=[130, 130, 130])
            tot_tbl.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#2A5ADF")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A5ADF")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 1), (-1, 1), "RIGHT"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ]))
            story.append(tot_tbl)

            # Construir PDF principal temporal (landscape)
            tmp_main = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            tmp_main_path = tmp_main.name
            tmp_main.close()

            doc = SimpleDocTemplate(
                tmp_main_path,
                pagesize=landscape(LETTER),
                leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36
            )
            doc.build(story)

            # Anexos (conduces)
            import shutil

            anexos = self._collect_conduces_to_attach()
            if not anexos:
                # Mover incluso entre discos diferentes
                shutil.move(tmp_main_path, out_path)
                return True, None

            annex_pdf_paths = []
            for a in anexos:
                if a["type"] == "pdf":
                    annex_pdf_paths.append(a["path"])
                else:
                    page_pdf = self._image_to_pdf_page(a["path"], a["label"])
                    if page_pdf:
                        annex_pdf_paths.append(page_pdf)

            ok, err = self._merge_main_with_annexes(tmp_main_path, annex_pdf_paths, out_path)
            if not ok:
                # Si falla el merge, al menos movemos el principal
                shutil.move(tmp_main_path, out_path)
                return False, f"No se pudieron anexar algunos conduces: {err}"

            # Limpieza de temporales (si tienes self._limpiar_temp_files)
            try:
                self._limpiar_temp_files()
            except Exception:
                pass

            return True, None

        except Exception as e:
            import traceback
            print("to_pdf error:", e)
            traceback.print_exc()
            return False, str(e) 
 
    def _agregar_anexos_conduces(self, elementos, estilos):
        """
        Agrega una sección de anexos con las imágenes de los conduces.
        
        Args:
            elementos: Lista de elementos del PDF
            estilos: Estilos de ReportLab
        """
        try:
            # Filtrar registros que tengan conduce
            conduces_df = self.df[self.df['CondStorage'].notna() & (self.df['CondStorage'] != '')]
            
            if conduces_df.empty:
                logger.info("No hay conduces para agregar a los anexos")
                return
            
            logger.info(f"Agregando {len(conduces_df)} conduces a los anexos")
            
            # Nueva página para anexos
            elementos.append(PageBreak())
            elementos.append(Paragraph("<b>ANEXOS: Conduces de Servicios</b>", estilos['Heading1']))
            elementos.append(Spacer(1, 5*mm))
            
            # Descargar y agregar cada conduce
            for idx, row in conduces_df.iterrows():
                storage_path = row['CondStorage']
                fecha = row.get('Fecha', '')
                conduce_num = row.get('Conduce', f'Conduce {idx+1}')
                
                # Descargar conduce desde Storage
                temp_path = self._descargar_conduce(storage_path)
                
                if temp_path and os.path.exists(temp_path):
                    # Agregar etiqueta
                    elementos.append(Paragraph(
                        f"<b>Conduce:</b> {conduce_num} | <b>Fecha:</b> {fecha}",
                        estilos['Normal']
                    ))
                    elementos.append(Spacer(1, 2*mm))
                    
                    # Agregar imagen (máximo 180mm de ancho)
                    try:
                        img = Image(temp_path)
                        img._restrictSize(180*mm, 250*mm)  # Máximo ancho y alto
                        elementos.append(img)
                    except Exception as e:
                        logger.warning(f"No se pudo insertar imagen {storage_path}: {e}")
                        elementos.append(Paragraph(
                            f"<i>No se pudo cargar la imagen del conduce</i>",
                            estilos['Normal']
                        ))
                    
                    elementos.append(Spacer(1, 5*mm))
                else:
                    logger.warning(f"No se pudo descargar conduce: {storage_path}")
            
        except Exception as e:
            logger.error(f"Error al agregar anexos de conduces: {e}", exc_info=True)
    
    def _descargar_conduce(self, storage_path):
        """
        Descarga un conduce desde Firebase Storage a un archivo temporal.
        
        Args:
            storage_path: Ruta del archivo en Storage
            
        Returns:
            str: Ruta del archivo temporal o None si falla
        """
        if not self.storage_manager:
            return None
        
        try:
            # Crear archivo temporal
            ext = os.path.splitext(storage_path)[1] or '.jpg'
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
            temp_path = temp_file.name
            temp_file.close()
            
            # Descargar desde Storage
            exito = self.storage_manager.descargar_conduce(storage_path, temp_path)
            
            if exito:
                self.temp_files.append(temp_path)
                return temp_path
            else:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                return None
                
        except Exception as e:
            logger.error(f"Error al descargar conduce {storage_path}: {e}")
            return None
    
    def _limpiar_temp_files(self):
        """Elimina archivos temporales descargados."""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    logger.debug(f"Archivo temporal eliminado: {temp_file}")
            except Exception as e:
                logger.warning(f"No se pudo eliminar archivo temporal {temp_file}: {e}")
        
        self.temp_files = []
    
    def _agregar_seccion_abonos(self, elementos, estilos):
        """
        Agrega sección de abonos y totales al PDF.
        """
        try:
            elementos.append(PageBreak())
            elementos.append(Paragraph("<b>ABONOS REGISTRADOS</b>", estilos['Heading2']))
            elementos.append(Spacer(1, 3*mm))
            
            # Tabla de abonos
            if self.abonos:
                abonos_cols = ['Fecha', 'Concepto', 'Monto']
                abonos_data = [abonos_cols]
                
                for abono in self.abonos:
                    fecha = abono.get('fecha', '')
                    concepto = abono.get('concepto', '') or abono.get('descripcion', 'Abono')
                    monto = float(abono.get('monto', 0))
                    
                    abonos_data.append([
                        str(fecha),
                        str(concepto),
                        f"{self.currency} {monto:,.2f}"
                    ])
                
                abonos_tabla = Table(abonos_data, colWidths=[40*mm, 90*mm, 50*mm])
                abonos_tabla.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F6321")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                    ('ALIGN', (0, 1), (1, -1), 'LEFT'),
                    
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")])
                ]))
                
                elementos.append(abonos_tabla)
                elementos.append(Spacer(1, 5*mm))
            else:
                elementos.append(Paragraph("No se registraron abonos en este período.", estilos['Normal']))
                elementos.append(Spacer(1, 5*mm))
            
            # Agregar totales
            self._agregar_totales(elementos, estilos)
            
        except Exception as e:
            logger.error(f"Error al agregar sección de abonos: {e}", exc_info=True)
    
    def _agregar_totales(self, elementos, estilos):
        """
        Agrega tabla de totales (Facturado, Abonado, Saldo).
        """
        try:
            elementos.append(Paragraph("<b>RESUMEN DE CUENTA</b>", estilos['Heading3']))
            elementos.append(Spacer(1, 3*mm))
            
            # Tabla de totales
            totales_data = [
                ['Total Facturado:', f"{self.currency} {self.total_facturado:,.2f}"],
                ['Total Abonado:', f"{self.currency} {self.total_abonado:,.2f}"],
                ['Saldo Pendiente:', f"{self.currency} {self.saldo:,.2f}"]
            ]
            
            totales_tabla = Table(totales_data, colWidths=[70*mm, 70*mm])
            totales_tabla.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                
                # Destacar el saldo
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor("#FFF4CC") if self.saldo > 0 else colors.HexColor("#E8F5E9")),
                ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor("#D32F2F") if self.saldo > 0 else colors.HexColor("#2E7D32")),
                
                ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
                ('LINEABOVE', (0, 2), (-1, 2), 2, colors.black),
                ('LINEBELOW', (0, 2), (-1, 2), 2, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elementos.append(totales_tabla)
            elementos.append(Spacer(1, 10*mm))
            
        except Exception as e:
            logger.error(f"Error al agregar totales: {e}", exc_info=True)
    
    def to_excel(self, filepath):
        """
        Genera el reporte en formato Excel.
        
        Args:
            filepath: Ruta donde guardar el Excel
            
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        if self.df.empty:
            return False, "No hay datos para exportar."
        
        try:
            # Exportar DataFrame a Excel
            self.df.to_excel(filepath, index=False, sheet_name='Reporte')
            logger.info(f"Excel generado exitosamente: {filepath}")
            return True, f"Reporte Excel generado exitosamente en:\n{filepath}"
            
        except Exception as e:
            logger.error(f"Error al generar Excel: {e}", exc_info=True)
            return False, f"Error al generar el reporte Excel: {str(e)}"


    def _column_widths_from_keys(self, column_map: dict) -> list:
        """
        Define anchos fijos por columna (en puntos). Evita que la tabla se distorsione.
        Ajusta estos valores si quieres más espacio en alguna columna.
        """
        width_map = {
            "fecha": 70,
            "conduce": 60,
            "ubicacion": 120,
            "equipo_nombre": 100,
            "operador_nombre": 120,
            "horas": 45,
            "monto": 85,
            "cliente_nombre": 120,
        }
        return [width_map.get(k, 80) for k in column_map.keys()]


    def _collect_conduces_to_attach(self) -> list[dict]:
        """
        Recorre self.data y devuelve una lista de anexos a agregar:
        [{"label": "Conduce No. 00621", "path": "/tmp/...", "type": "image|pdf"}]
        Descarga cada conduce con storage_manager.descargar_conduce(storage_path).
        Acepta las claves 'conduce_storage_path' o 'CondStorage' como ruta en Storage.
        """
        anexos = []
        sm = getattr(self, "storage_manager", None)
        if not sm:
            return anexos

        for row in self.data or []:
            storage_path = row.get("conduce_storage_path") or row.get("CondStorage")
            if not storage_path:
                continue

            # Si ya viene una URL http(s), no podemos usar descargar_conduce; opcionalmente podrías bajarlo con requests.
            # En tu app guardas storage_path (conduces/YYYY/MM/file.ext), así que debería funcionar.
            try:
                local_path = sm.descargar_conduce(storage_path)
                if not local_path:
                    continue
                import os  # si no lo tienes ya importado arriba
                ext = os.path.splitext(local_path)[1].lower()
                numero = str(row.get("conduce") or row.get("id") or "")
                label = f"Conduce No. {numero}" if numero else "Conduce"
                tipo = "pdf" if ext == ".pdf" else "image"
                anexos.append({"label": label, "path": local_path, "type": tipo})
            except Exception:
                continue

        return anexos


    def _image_to_pdf_page(self, img_path: str, label: str) -> str | None:
        """
        Convierte una imagen en una página PDF con un título (label) arriba.
        Devuelve la ruta al PDF temporal creado.
        """
        try:
            import tempfile
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import LETTER
            from reportlab.lib.utils import ImageReader

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            tmp_path = tmp.name
            tmp.close()

            page_w, page_h = LETTER
            margin = 36  # 0.5in
            title_space = 20

            c = canvas.Canvas(tmp_path, pagesize=LETTER)
            c.setFont("Helvetica-Bold", 12)
            c.drawString(margin, page_h - margin, label)

            # Cargar imagen y escalar manteniendo aspecto
            img = ImageReader(img_path)
            iw, ih = img.getSize()
            max_w = page_w - 2 * margin
            max_h = page_h - 2 * margin - title_space
            scale = min(max_w / iw, max_h / ih)
            draw_w = iw * scale
            draw_h = ih * scale
            x = (page_w - draw_w) / 2
            y = margin
            c.drawImage(img, x, y, draw_w, draw_h, preserveAspectRatio=True, anchor='sw')

            c.showPage()
            c.save()
            return tmp_path
        except Exception:
            return None


    def _merge_main_with_annexes(self, main_pdf: str, annex_pdf_paths: list[str], out_path: str) -> tuple[bool, str | None]:
        """
        Fusiona el PDF principal con una lista de PDFs anexos al final.
        """
        try:
            from PyPDF2 import PdfMerger
            merger = PdfMerger()
            merger.append(main_pdf)
            for p in annex_pdf_paths:
                merger.append(p)
            with open(out_path, "wb") as f:
                merger.write(f)
            merger.close()
            return True, None
        except Exception as e:
            return False, str(e)
        
    def _scale_col_widths_to_page(self, col_widths: list, page_width: float, left_margin: float, right_margin: float) -> list:
        """
        Si la suma de colWidths excede el ancho disponible en la página, escala proporcionalmente para que quepan.
        """
        if not col_widths:
            return col_widths
        available = page_width - (left_margin + right_margin)
        total = sum(col_widths)
        if total <= 0 or total <= available:
            return col_widths
        scale = available / total
        return [w * scale for w in col_widths]
    
    def _is_flexible_col(self, key: str) -> bool:
        """
        Columnas que pueden envolver el texto en varias líneas si no caben.
        """
        return key in {"ubicacion", "equipo_nombre", "operador_nombre", "cliente_nombre"}


    def _format_value_for_measure(self, key: str, val) -> str:
        """
        Devuelve el texto final como se mostrará en la celda (para medir y/o renderizar).
        """
        if key == "horas" and val not in ("", None):
            try:
                return f"{float(val):,.2f}"
            except Exception:
                return str(val or "")
        if key == "monto" and val not in ("", None):
            try:
                return f"{self.currency_symbol} {float(val):,.2f}"
            except Exception:
                return str(val or "")
        return str(val or "")


    def _auto_compute_col_widths(self, column_map: dict, data: list, page_w: float, margins=(36, 36),
                                font_name="Helvetica", font_size=9) -> list[float]:
        """
        Calcula anchos por columna según el texto más largo (headers + filas).
        Si excede el ancho disponible, reduce proporcionalmente SOLO columnas flexibles,
        dejando que el texto se envuelva (las filas aumentan de alto).
        """
        from reportlab.pdfbase.pdfmetrics import stringWidth

        left_margin, right_margin = margins
        available = page_w - (left_margin + right_margin)
        keys = list(column_map.keys())

        # Mínimos por columna (nunca bajamos de esto)
        min_widths = {
            "fecha": 60.0,
            "conduce": 55.0,
            "ubicacion": 90.0,
            "equipo_nombre": 90.0,
            "operador_nombre": 110.0,
            "horas": 45.0,
            "monto": 85.0,
            "cliente_nombre": 110.0,
        }
        pad = 10.0  # padding horizontal estimado por celda

        # 1) Deseados por medición (header + filas)
        desired = []
        for k in keys:
            header_text = str(column_map[k])
            max_w = stringWidth(header_text, font_name, font_size)
            for r in data:
                txt = self._format_value_for_measure(k, r.get(k, ""))
                w = stringWidth(txt, font_name, font_size)
                if w > max_w:
                    max_w = w
            desired_w = max_w + pad
            desired.append(desired_w)

        # 2) Aplicar mínimos
        widths = []
        for k, w in zip(keys, desired):
            widths.append(max(w, min_widths.get(k, 70.0)))

        total = sum(widths)
        if total <= available or available <= 0:
            return widths

        # 3) Reducir solo columnas flexibles hasta caber (proporcional a su "exceso" sobre el mínimo)
        for _ in range(3):  # hasta 3 iteraciones por si queda residuo
            total = sum(widths)
            if total <= available:
                break
            over = total - available

            # Espacio reducible en flex (por encima del mínimo)
            flex_excess = 0.0
            for i, k in enumerate(keys):
                if self._is_flexible_col(k):
                    flex_excess += max(0.0, widths[i] - min_widths.get(k, 70.0))

            if flex_excess <= 1e-6:
                # No hay de dónde reducir (ya estamos en mínimos) -> devolvemos widths, el wrapping hará el resto
                return widths

            # Reducir proporcionalmente a su exceso
            new_widths = list(widths)
            for i, k in enumerate(keys):
                if not self._is_flexible_col(k):
                    continue
                exceso_col = max(0.0, widths[i] - min_widths.get(k, 70.0))
                if exceso_col <= 0:
                    continue
                reduce_i = over * (exceso_col / flex_excess)
                new_w = max(min_widths.get(k, 70.0), widths[i] - reduce_i)
                new_widths[i] = new_w
            widths = new_widths

        return widths


    def _make_wrap_paragraph(self, text: str, font_size=9, align=TA_LEFT) -> Paragraph:
        """
        Crea un Paragraph con wrap para celdas de texto.
        """
        from reportlab.lib.styles import ParagraphStyle
        style = ParagraphStyle(
            name="Cell",
            parent=getSampleStyleSheet()["BodyText"],
            fontName="Helvetica",
            fontSize=font_size,
            leading=font_size + 2,
            alignment=align,
        )
        # escape para caracteres especiales HTML
        return Paragraph(escape(text or ""), style)


    def _rows_with_wrapping(self, column_map: dict, data: list, font_size=9) -> list[list]:
        """
        Construye filas: para columnas flexibles usa Paragraph (wrap) y en numéricas aplica formato.
        """
        rows = []
        keys = list(column_map.keys())
        for r in data:
            row_cells = []
            for k in keys:
                txt = self._format_value_for_measure(k, r.get(k, ""))
                if self._is_flexible_col(k):
                    row_cells.append(self._make_wrap_paragraph(txt, font_size=font_size))
                else:
                    row_cells.append(txt)
            rows.append(row_cells)
        return rows
    
    def generar_reporte_rendimientos_bloques(
        self,
        file_path: str,
        formato: str,
        datos_facturacion: list[dict],
        datos_rendimientos: list[dict],
        resumen: dict,
        moneda: str = "RD$",
        titulo: str = "REPORTE DE RENDIMIENTOS",
        rango_fechas:  str = "",
    ) -> tuple[bool, str | None]:
        """
        Genera reporte de rendimientos con estructura de bloques. 
        
        Soporta PDF y Excel con 3 secciones: 
        1. Bloque Facturación
        2. Bloque Rendimientos
        3. Resumen General
        
        Args:
            file_path:  Ruta donde guardar el archivo
            formato:  "pdf" o "excel"
            datos_facturacion: Lista de dicts con datos de facturación
            datos_rendimientos: Lista de dicts con datos de rendimientos
            resumen: Dict con totales generales
            moneda: Símbolo de moneda
            titulo: Título del reporte
            rango_fechas: Rango de fechas (ej: "2025-01-01 a 2025-01-31")
        
        Returns:
            tuple:  (éxito:  bool, error: str | None)
        """
        try: 
            if formato == "pdf":
                return self._generar_pdf_rendimientos_bloques(
                    file_path, datos_facturacion, datos_rendimientos,
                    resumen, moneda, titulo, rango_fechas
                )
            elif formato == "excel":
                return self._generar_excel_rendimientos_bloques(
                    file_path, datos_facturacion, datos_rendimientos,
                    resumen, moneda, titulo, rango_fechas
                )
            else:
                return False, f"Formato no soportado:  {formato}"
        
        except Exception as e:
            logger.error(f"Error generando reporte rendimientos: {e}", exc_info=True)
            return False, str(e)

    def generar_reporte_rendimientos_bloques(
        self,
        file_path: str,
        formato: str,
        datos_facturacion: list[dict],
        datos_rendimientos: list[dict],
        resumen: dict,
        moneda: str = "RD$",
        titulo: str = "REPORTE DE RENDIMIENTOS",
        rango_fechas: str = "",
    ) -> tuple[bool, str | None]:
        """
        Genera reporte de rendimientos con estructura de bloques. 
        
        Soporta PDF y Excel con 3 secciones: 
        1. Bloque Facturación
        2. Bloque Rendimientos
        3. Resumen General
        
        Args:
            file_path:  Ruta donde guardar el archivo
            formato:  "pdf" o "excel"
            datos_facturacion: Lista de dicts con datos de facturación
            datos_rendimientos: Lista de dicts con datos de rendimientos
            resumen: Dict con totales generales
            moneda: Símbolo de moneda
            titulo: Título del reporte
            rango_fechas: Rango de fechas (ej: "2025-01-01 a 2025-01-31")
        
        Returns:
            tuple:  (éxito:  bool, error: str | None)
        """
        try: 
            if formato == "pdf":
                return self._generar_pdf_rendimientos_bloques(
                    file_path, datos_facturacion, datos_rendimientos,
                    resumen, moneda, titulo, rango_fechas
                )
            elif formato == "excel":
                return self._generar_excel_rendimientos_bloques(
                    file_path, datos_facturacion, datos_rendimientos,
                    resumen, moneda, titulo, rango_fechas
                )
            else:
                return False, f"Formato no soportado:  {formato}"
        
        except Exception as e:
            logger.error(f"Error generando reporte rendimientos: {e}", exc_info=True)
            return False, str(e)

    def _generar_pdf_rendimientos_bloques(
        self,
        file_path: str,
        datos_facturacion: list,
        datos_rendimientos: list,
        resumen: dict,
        moneda: str,
        titulo: str,
        rango_fechas: str,
    ) -> tuple[bool, str | None]:
        """Genera PDF con bloques de Facturación + Rendimientos + Resumen."""
        try:
            print("\n=== INICIO _generar_pdf_rendimientos_bloques ===")
            print(f"File path: {file_path}")
            print(f"Moneda: {moneda}")
            print(f"Título: {titulo}")
            print(f"Rango:  {rango_fechas}")
            print(f"Datos facturación: {len(datos_facturacion)}")
            print(f"Datos rendimientos: {len(datos_rendimientos)}")
            print(f"Resumen: {resumen}")
            
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            from reportlab.lib import colors

            # Crear documento
            print("\n1. Creando documento...")
            doc = SimpleDocTemplate(
                file_path,
                pagesize=landscape(letter),
                leftMargin=36,
                rightMargin=36,
                topMargin=50,
                bottomMargin=50,
            )

            story = []
            styles = getSampleStyleSheet()

            # Estilo personalizado
            style_seccion = ParagraphStyle(
                name="Seccion",
                parent=styles["Heading2"],
                fontSize=12,
                textColor=colors.HexColor("#1F7A1F"),
                spaceAfter=8,
                fontName="Helvetica-Bold",
            )

            # --- ENCABEZADO ---
            print("\n2. Agregando encabezado...")
            story.append(Paragraph(f"<b>{titulo}</b>", styles["Title"]))
            if rango_fechas:
                story.append(Paragraph(f"Período: {rango_fechas}", styles["Normal"]))
            story.append(Spacer(1, 15))

            # --- BLOQUE 1: FACTURACIÓN ---
            print("\n3. Procesando BLOQUE FACTURACIÓN...")
            story.append(Paragraph("📊 FACTURACIÓN", style_seccion))
            
            if datos_facturacion:
                headers_fact = [
                    "Equipo", "Horas Fact.", "Volumen Fact.", "Monto Facturado",
                    "Precio/h", "Precio/u", "Modalidad(s)"
                ]
                
                data_fact = [headers_fact]
                
                for i, d in enumerate(datos_facturacion):
                    print(f"\n  Fila {i} facturación:")
                    print(f"    Datos: {d}")
                    
                    try:
                        equipo = str(d.get("equipo", ""))
                        print(f"    Equipo:  {equipo}")
                        
                        horas = float(d.get('horas_facturadas', 0))
                        horas_txt = f"{horas:.2f} h"
                        print(f"    Horas:  {horas} -> {horas_txt}")
                        
                        volumen = float(d.get('volumen_facturado', 0))
                        volumen_txt = f"{volumen:.2f}"
                        print(f"    Volumen: {volumen} -> {volumen_txt}")
                        
                        monto = float(d.get('monto_facturado', 0))
                        monto_txt = f"{moneda} {monto:,.2f}"
                        print(f"    Monto:  {monto} -> {monto_txt}")
                        
                        precio_h = float(d.get('precio_hora_facturado', 0))
                        precio_h_txt = f"{moneda} {precio_h:,.2f}"
                        print(f"    Precio/h: {precio_h} -> {precio_h_txt}")
                        
                        precio_u = float(d.get('precio_unidad_facturado', 0))
                        precio_u_txt = f"{moneda} {precio_u:,.2f}"
                        print(f"    Precio/u:  {precio_u} -> {precio_u_txt}")
                        
                        modalidades = str(d.get("modalidades", "-"))
                        print(f"    Modalidades: {modalidades}")
                        
                        fila = [equipo, horas_txt, volumen_txt, monto_txt, precio_h_txt, precio_u_txt, modalidades]
                        print(f"    Fila completa: {fila}")
                        data_fact.append(fila)
                        
                    except Exception as e: 
                        print(f"    ❌ ERROR en fila {i}: {e}")
                        import traceback
                        traceback.print_exc()
                        raise
                
                print(f"\n  Total filas facturación: {len(data_fact)}")
                print("  Creando tabla facturación...")
                
                tabla_fact = Table(data_fact, hAlign="LEFT")
                tabla_fact.setStyle(TableStyle([
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6F4EA")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F7A1F")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1F7A1F")),
                    ("ALIGN", (1, 1), (5, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
                ]))
                story.append(tabla_fact)
                print("  ✓ Tabla facturación creada")
            else:
                story.append(Paragraph("Sin datos de facturación", styles["Normal"]))
            
            story.append(Spacer(1, 20))

            # --- BLOQUE 2: RENDIMIENTOS ---
            print("\n4. Procesando BLOQUE RENDIMIENTOS...")
            story.append(Paragraph("💰 RENDIMIENTOS", style_seccion))
            
            if datos_rendimientos:
                headers_rend = [
                    "Equipo", "Horas Pag.", "Pagado Op.", "Gastos Equipo",
                    "Rendimiento Neto", "% Margen"
                ]
                
                data_rend = [headers_rend]
                
                for i, d in enumerate(datos_rendimientos):
                    print(f"\n  Fila {i} rendimientos:")
                    print(f"    Datos: {d}")
                    
                    try:
                        equipo = str(d.get("equipo", ""))
                        horas_pag = float(d.get('horas_pagadas', 0))
                        pag_op = float(d.get('monto_pagado_operador', 0))
                        gastos = float(d.get('gastos_equipo', 0))
                        rend_neto = float(d.get('rendimiento_neto', 0))
                        margen = float(d.get('margen_porcentaje', 0))
                        
                        fila = [
                            equipo,
                            f"{horas_pag:.2f} h",
                            f"{moneda} {pag_op:,.2f}",
                            f"{moneda} {gastos:,.2f}",
                            f"{moneda} {rend_neto:,.2f}",
                            f"{margen:.2f}%",
                        ]
                        print(f"    Fila:  {fila}")
                        data_rend.append(fila)
                        
                    except Exception as e:
                        print(f"    ❌ ERROR en fila {i}: {e}")
                        import traceback
                        traceback.print_exc()
                        raise
                
                print("  Creando tabla rendimientos...")
                tabla_rend = Table(data_rend, hAlign="LEFT")
                tabla_rend.setStyle(TableStyle([
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF4E0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#A35D00")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A35D00")),
                    ("ALIGN", (1, 1), (5, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFFAF0")]),
                ]))
                story.append(tabla_rend)
                print("  ✓ Tabla rendimientos creada")
            else:
                story.append(Paragraph("Sin datos de rendimientos", styles["Normal"]))
            
            story.append(Spacer(1, 25))

            # --- RESUMEN GENERAL ---
            print("\n5. Procesando RESUMEN...")
            print(f"  Datos resumen: {resumen}")
            story.append(Paragraph("📈 RESUMEN GENERAL", style_seccion))
            
            try:
                total_horas = float(resumen.get('total_horas_facturadas', 0))
                total_fact = float(resumen.get('total_facturado', 0))
                total_pag = float(resumen.get('total_pagado_operador', 0))
                total_gast = float(resumen.get('total_gastos', 0))
                rend_neto = float(resumen.get('rendimiento_neto', 0))
                margen_prom = float(resumen.get('margen_promedio', 0))
                
                print(f"  Valores extraídos:")
                print(f"    Horas:  {total_horas}")
                print(f"    Facturado: {total_fact}")
                print(f"    Pagado:  {total_pag}")
                print(f"    Gastos: {total_gast}")
                print(f"    Rendimiento: {rend_neto}")
                print(f"    Margen: {margen_prom}")
                
                resumen_data = [
                    ["Total Horas Facturadas:", f"{total_horas:,.2f} h"],
                    ["Total Facturado:", f"{moneda} {total_fact:,.2f}"],
                    ["Total Pagado Operador:", f"{moneda} {total_pag:,.2f}"],
                    ["Total Gastos Equipos:", f"{moneda} {total_gast:,.2f}"],
                    ["", ""],
                    ["Rendimiento Neto:", f"{moneda} {rend_neto:,.2f}"],
                    ["Margen Promedio:", f"{margen_prom:.2f}%"],
                ]
                
                print("  Creando tabla resumen...")
                tabla_resumen = Table(resumen_data, colWidths=[200, 150], hAlign="CENTER")
                tabla_resumen.setStyle(TableStyle([
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("LINEABOVE", (0, 5), (-1, 5), 2, colors.HexColor("#2E7D32")),
                    ("BACKGROUND", (0, 5), (-1, 6), colors.HexColor("#E8F5E9")),
                    ("TEXTCOLOR", (0, 5), (-1, 6), colors.HexColor("#2E7D32")),
                    ("FONTNAME", (0, 5), (-1, 6), "Helvetica-Bold"),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]))
                story.append(tabla_resumen)
                print("  ✓ Tabla resumen creada")
                
            except Exception as e:
                print(f"  ❌ ERROR creando resumen: {e}")
                import traceback
                traceback.print_exc()
                raise

            # Construir PDF
            print("\n6. Construyendo PDF...")
            doc.build(story)
            print("  ✓ PDF construido exitosamente")
            
            print("=== FIN _generar_pdf_rendimientos_bloques ===\n")
            return True, None

        except Exception as e: 
            print(f"\n❌ ERROR GENERAL en _generar_pdf_rendimientos_bloques: {e}")
            import traceback
            traceback.print_exc()
            logger.error(f"Error generando PDF rendimientos: {e}", exc_info=True)
            return False, str(e)        

    def _generar_excel_rendimientos_bloques(
        self,
        file_path: str,
        datos_facturacion: list,
        datos_rendimientos: list,
        resumen: dict,
        moneda: str,
        titulo: str,
        rango_fechas: str,
    ) -> tuple[bool, str | None]:
        """
        Genera Excel con 3 hojas: 
        - Hoja 1: Facturación
        - Hoja 2: Rendimientos
        - Hoja 3: Resumen
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Crear DataFrames
            df_fact = pd.DataFrame(datos_facturacion) if datos_facturacion else pd.DataFrame()
            df_rend = pd.DataFrame(datos_rendimientos) if datos_rendimientos else pd.DataFrame()

            # Formatear columnas de facturación
            if not df_fact.empty:
                df_fact = df_fact.rename(columns={
                    "equipo": "Equipo",
                    "horas_facturadas": "Horas Fact.",
                    "volumen_facturado": "Volumen Fact.",
                    "monto_facturado": "Monto Facturado",
                    "precio_hora_facturado": "Precio/h",
                    "precio_unidad_facturado":  "Precio/u",
                    "modalidades": "Modalidad(s)",
                })

            # Formatear columnas de rendimientos
            if not df_rend.empty:
                df_rend = df_rend.rename(columns={
                    "equipo": "Equipo",
                    "horas_pagadas": "Horas Pag.",
                    "monto_pagado_operador": "Pagado Op.",
                    "gastos_equipo": "Gastos Equipo",
                    "rendimiento_neto": "Rendimiento Neto",
                    "margen_porcentaje": "% Margen",
                })

            # Escribir a Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Hoja 1: Facturación
                if not df_fact.empty:
                    df_fact.to_excel(writer, sheet_name='Facturación', index=False)
                
                # Hoja 2: Rendimientos
                if not df_rend.empty:
                    df_rend.to_excel(writer, sheet_name='Rendimientos', index=False)
                
                # Hoja 3: Resumen
                resumen_data = {
                    "Concepto": [
                        "Total Horas Facturadas",
                        "Total Facturado",
                        "Total Pagado Operador",
                        "Total Gastos Equipos",
                        "",
                        "Rendimiento Neto",
                        "Margen Promedio",
                    ],
                    "Valor": [
                        f"{resumen.get('total_horas_facturadas', 0):,.2f} h",
                        f"{moneda} {resumen.get('total_facturado', 0):,.2f}",
                        f"{moneda} {resumen.get('total_pagado_operador', 0):,.2f}",
                        f"{moneda} {resumen.get('total_gastos', 0):,.2f}",
                        "",
                        f"{moneda} {resumen.get('rendimiento_neto', 0):,.2f}",
                        f"{resumen.get('margen_promedio', 0):,.2f}%",
                    ],
                }
                df_resumen = pd.DataFrame(resumen_data)
                df_resumen.to_excel(writer, sheet_name='Resumen', index=False)

            # Aplicar estilos con openpyxl
            wb = load_workbook(file_path)

            # Estilos comunes
            header_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            header_font = Font(bold=True, color="1F7A1F")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Formatear hoja Facturación
            if 'Facturación' in wb.sheetnames:
                ws = wb['Facturación']
                for cell in ws[1]:  # Header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar anchos
                ws.column_dimensions['A'].width = 25  # Equipo
                ws.column_dimensions['B'].width = 12  # Horas
                ws.column_dimensions['C'].width = 12  # Volumen
                ws.column_dimensions['D'].width = 15  # Monto
                ws.column_dimensions['E'].width = 12  # Precio/h
                ws.column_dimensions['F'].width = 12  # Precio/u
                ws.column_dimensions['G'].width = 15  # Modalidad

            # Formatear hoja Rendimientos
            if 'Rendimientos' in wb.sheetnames:
                ws = wb['Rendimientos']
                for cell in ws[1]: 
                    cell.fill = PatternFill(start_color="FFF4E0", end_color="FFF4E0", fill_type="solid")
                    cell.font = Font(bold=True, color="A35D00")
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 18
                ws.column_dimensions['F'].width = 12

            # Formatear hoja Resumen
            if 'Resumen' in wb.sheetnames:
                ws = wb['Resumen']
                
                # Header
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    cell.font = Font(bold=True, color="2E7D32")
                    cell.border = border
                
                # Destacar Rendimiento Neto y Margen
                for row_idx in [7, 8]:  # Filas de rendimiento y margen
                    for cell in ws[row_idx]:
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                        cell.font = Font(bold=True, color="2E7D32")
                
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20

            # Guardar cambios
            wb.save(file_path)
            
            return True, None

        except Exception as e:
            logger.error(f"Error generando Excel rendimientos: {e}", exc_info=True)
            return False, str(e)