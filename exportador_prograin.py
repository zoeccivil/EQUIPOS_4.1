"""
Exportador de transacciones al formato compatible con PROGRAIN 5.0
Genera archivos Excel (.xlsx) con el formato requerido para importación
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers

logger = logging.getLogger(__name__)


class ExportadorPrograin:
    """Exportador de transacciones al formato compatible con PROGRAIN 5.0"""
    
    def __init__(self, moneda_symbol: str = "RD$"):
        """Inicializa el exportador con el símbolo de moneda"""
        self.moneda_symbol = moneda_symbol
    
    def exportar_transacciones(
        self,
        gastos: List[Dict[str, Any]],
        ingresos: List[Dict[str, Any]],
        output_path: str,
        mapas: Dict[str, Dict[str, str]],
        incluir_gastos: bool = True,
        incluir_ingresos: bool = True
    ) -> bool:
        """
        Exporta transacciones al formato PROGRAIN 5.0
        
        Args:
            gastos: Lista de gastos desde Firebase (colección 'gastos')
            ingresos: Lista de ingresos desde Firebase (colección 'alquileres')
            output_path: Ruta completa del archivo .xlsx a crear
            mapas: Dict con mapas de nombres (equipos, clientes, cuentas, categorías, subcategorías, proyectos)
            incluir_gastos: Si True, incluye gastos en la exportación
            incluir_ingresos: Si True, incluye ingresos en la exportación
        
        Returns:
            True si la exportación fue exitosa, False si hubo error
        """
        try:
            transacciones = []
            
            # Convertir gastos
            if incluir_gastos:
                for gasto in gastos:
                    trans = self._convertir_gasto_a_transaccion(gasto, mapas)
                    if trans:
                        transacciones.append(trans)
                logger.info(f"Convertidos {len([t for t in transacciones if t['Débito'] > 0])} gastos")
            
            # Convertir ingresos
            if incluir_ingresos:
                for ingreso in ingresos:
                    trans = self._convertir_ingreso_a_transaccion(ingreso, mapas)
                    if trans:
                        transacciones.append(trans)
                logger.info(f"Convertidos {len([t for t in transacciones if t['Crédito'] > 0])} ingresos")
            
            if not transacciones:
                logger.warning("No hay transacciones para exportar")
                return False
            
            # Crear DataFrame
            df = pd.DataFrame(transacciones)
            
            # Ordenar por fecha ascendente
            df = df.sort_values('Fecha')
            
            # Convertir fecha a datetime nativo
            df['Fecha'] = pd.to_datetime(df['Fecha'])
            
            # Asegurar que montos sean float
            df['Débito'] = df['Débito'].astype(float)
            df['Crédito'] = df['Crédito'].astype(float)
            
            # Exportar a Excel
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Ajustar formato visual
            self._ajustar_formato_columnas(output_path)
            
            logger.info(f"Exportación exitosa: {len(transacciones)} transacciones en {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando transacciones: {e}", exc_info=True)
            return False
    
    def _convertir_gasto_a_transaccion(self, gasto: Dict, mapas: Dict) -> Dict:
        """
        Convierte un gasto de Firebase al formato PROGRAIN
        
        Mapeo:
        - Fecha: gasto['fecha']
        - Concepto: Nombre de la categoría (mapas['categorias'][gasto['categoria_id']])
        - Detalle: "[{equipo_nombre}] {descripcion} ({comentario})"
        - Débito: gasto['monto'] (como float)
        - Crédito: 0.00
        """
        try:
            # Obtener fecha
            fecha = gasto.get('fecha', '')
            if not fecha:
                logger.warning(f"Gasto {gasto.get('id')} sin fecha, se omite")
                return None
            
            # Obtener concepto (categoría)
            categoria_id = str(gasto.get('categoria_id', ''))
            concepto = mapas.get('categorias', {}).get(categoria_id, 'GASTO SIN CATEGORÍA')
            
            # Construir detalle
            partes_detalle = []
            
            # Equipo
            equipo_id = gasto.get('equipo_id')
            if equipo_id:
                equipo_nombre = mapas.get('equipos', {}).get(str(equipo_id), f'Equipo {equipo_id}')
                partes_detalle.append(f"[{equipo_nombre}]")
            
            # Descripción
            descripcion = gasto.get('descripcion', '').strip()
            if descripcion:
                partes_detalle.append(descripcion)
            
            # Comentario
            comentario = gasto.get('comentario', '').strip()
            if comentario:
                partes_detalle.append(f"({comentario})")
            
            detalle = ' '.join(partes_detalle) if partes_detalle else concepto
            
            # Limitar longitud del detalle
            if len(detalle) > 200:
                detalle = detalle[:197] + '...'
            
            # Obtener monto
            monto = float(gasto.get('monto', 0))
            if monto <= 0:
                logger.warning(f"Gasto {gasto.get('id')} con monto <= 0, se omite")
                return None
            
            return {
                'Fecha': fecha,
                'Concepto': concepto[:200],  # Limitar a 200 caracteres
                'Detalle': detalle,
                'Débito': monto,
                'Crédito': 0.00
            }
            
        except Exception as e:
            logger.error(f"Error convirtiendo gasto {gasto.get('id')}: {e}", exc_info=True)
            return None
    
    def _convertir_ingreso_a_transaccion(self, ingreso: Dict, mapas: Dict) -> Dict:
        """
        Convierte un ingreso (alquiler) de Firebase al formato PROGRAIN
        
        Mapeo:
        - Fecha: ingreso['fecha']
        - Concepto: "INGRESO ALQUILER" (fijo)
        - Detalle: "{equipo_nombre} - {cliente_nombre} - {proyecto_nombre}"
        - Débito: 0.00
        - Crédito: ingreso['monto'] (como float)
        """
        try:
            # Obtener fecha
            fecha = ingreso.get('fecha', '')
            if not fecha:
                logger.warning(f"Ingreso {ingreso.get('id')} sin fecha, se omite")
                return None
            
            # Concepto fijo
            concepto = "INGRESO ALQUILER"
            
            # Construir detalle
            partes_detalle = []
            
            # Equipo
            equipo_id = ingreso.get('equipo_id')
            if equipo_id:
                equipo_nombre = mapas.get('equipos', {}).get(str(equipo_id), f'Equipo {equipo_id}')
                partes_detalle.append(equipo_nombre)
            
            # Cliente
            cliente_id = ingreso.get('cliente_id')
            if cliente_id:
                cliente_nombre = mapas.get('clientes', {}).get(str(cliente_id), f'Cliente {cliente_id}')
                partes_detalle.append(cliente_nombre)
            
            # Proyecto
            proyecto_id = ingreso.get('proyecto_id')
            if proyecto_id:
                proyecto_nombre = mapas.get('proyectos', {}).get(str(proyecto_id), f'Proyecto {proyecto_id}')
                partes_detalle.append(proyecto_nombre)
            
            detalle = ' - '.join(partes_detalle) if partes_detalle else 'INGRESO ALQUILER'
            
            # Limitar longitud del detalle
            if len(detalle) > 200:
                detalle = detalle[:197] + '...'
            
            # Obtener monto
            monto = float(ingreso.get('monto', 0))
            if monto <= 0:
                logger.warning(f"Ingreso {ingreso.get('id')} con monto <= 0, se omite")
                return None
            
            return {
                'Fecha': fecha,
                'Concepto': concepto,
                'Detalle': detalle,
                'Débito': 0.00,
                'Crédito': monto
            }
            
        except Exception as e:
            logger.error(f"Error convirtiendo ingreso {ingreso.get('id')}: {e}", exc_info=True)
            return None
    
    def _ajustar_formato_columnas(self, archivo_path: str):
        """
        Ajusta formato visual del Excel usando openpyxl:
        - Encabezados: negrita, fondo azul (#366092), texto blanco
        - Anchos de columnas: Fecha=12, Concepto=25, Detalle=50, Débito=15, Crédito=15
        - Formato de montos: separador de miles, 2 decimales, alineación derecha
        - Formato de fechas: YYYY-MM-DD, alineación centrada
        """
        try:
            wb = load_workbook(archivo_path)
            ws = wb.active
            
            # Estilos para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar estilos a encabezados (fila 1)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajustar anchos de columnas
            ws.column_dimensions['A'].width = 12  # Fecha
            ws.column_dimensions['B'].width = 25  # Concepto
            ws.column_dimensions['C'].width = 50  # Detalle
            ws.column_dimensions['D'].width = 15  # Débito
            ws.column_dimensions['E'].width = 15  # Crédito
            
            # Formato de celdas de datos (desde fila 2)
            for row in range(2, ws.max_row + 1):
                # Fecha: centrada, formato fecha
                fecha_cell = ws.cell(row=row, column=1)
                fecha_cell.alignment = Alignment(horizontal="center")
                fecha_cell.number_format = 'yyyy-mm-dd'
                
                # Concepto: alineación izquierda
                concepto_cell = ws.cell(row=row, column=2)
                concepto_cell.alignment = Alignment(horizontal="left")
                
                # Detalle: alineación izquierda
                detalle_cell = ws.cell(row=row, column=3)
                detalle_cell.alignment = Alignment(horizontal="left")
                
                # Débito: derecha, formato numérico
                debito_cell = ws.cell(row=row, column=4)
                debito_cell.alignment = Alignment(horizontal="right")
                debito_cell.number_format = '#,##0.00'
                
                # Crédito: derecha, formato numérico
                credito_cell = ws.cell(row=row, column=5)
                credito_cell.alignment = Alignment(horizontal="right")
                credito_cell.number_format = '#,##0.00'
            
            wb.save(archivo_path)
            logger.info(f"Formato de columnas ajustado en {archivo_path}")
            
        except Exception as e:
            logger.error(f"Error ajustando formato de columnas: {e}", exc_info=True)
    
    def validar_archivo_prograin(self, archivo_path: str) -> Dict[str, Any]:
        """
        Valida que un archivo Excel cumpla con el formato PROGRAIN 5.0
        
        Returns:
            {
                'valido': bool,
                'errores': List[str],
                'advertencias': List[str],
                'estadisticas': {
                    'total_transacciones': int,
                    'total_debitos': float,
                    'total_creditos': float,
                    'fecha_inicio': str,
                    'fecha_fin': str,
                    'balance': float  # Créditos - Débitos
                }
            }
        """
        resultado = {
            'valido': True,
            'errores': [],
            'advertencias': [],
            'estadisticas': {
                'total_transacciones': 0,
                'total_debitos': 0.0,
                'total_creditos': 0.0,
                'fecha_inicio': '',
                'fecha_fin': '',
                'balance': 0.0
            }
        }
        
        try:
            # Leer archivo
            df = pd.read_excel(archivo_path)
            
            # Validar columnas requeridas
            columnas_requeridas = ['Fecha', 'Concepto', 'Detalle', 'Débito', 'Crédito']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                resultado['errores'].append(f"Faltan columnas requeridas: {', '.join(columnas_faltantes)}")
                resultado['valido'] = False
                return resultado
            
            # Validar que haya datos
            if len(df) == 0:
                resultado['errores'].append("El archivo no contiene transacciones")
                resultado['valido'] = False
                return resultado
            
            # Validar tipos de datos
            errores_fecha = []
            for idx, fecha in enumerate(df['Fecha']):
                try:
                    pd.to_datetime(fecha)
                except:
                    errores_fecha.append(f"Fila {idx + 2}: fecha inválida '{fecha}'")
            
            if errores_fecha:
                resultado['errores'].extend(errores_fecha[:5])  # Mostrar máximo 5
                if len(errores_fecha) > 5:
                    resultado['errores'].append(f"... y {len(errores_fecha) - 5} errores más de fecha")
                resultado['valido'] = False
            
            # Validar montos
            for idx, row in df.iterrows():
                try:
                    debito = float(row['Débito'])
                    credito = float(row['Crédito'])
                    
                    # Validar que no sean negativos
                    if debito < 0:
                        resultado['errores'].append(f"Fila {idx + 2}: Débito negativo ({debito})")
                        resultado['valido'] = False
                    
                    if credito < 0:
                        resultado['errores'].append(f"Fila {idx + 2}: Crédito negativo ({credito})")
                        resultado['valido'] = False
                    
                    # Validar que al menos uno sea 0
                    if debito > 0 and credito > 0:
                        resultado['advertencias'].append(
                            f"Fila {idx + 2}: Tanto Débito como Crédito tienen valores (se esperaba solo uno)"
                        )
                    
                    # Validar que al menos uno tenga valor
                    if debito == 0 and credito == 0:
                        resultado['advertencias'].append(
                            f"Fila {idx + 2}: Tanto Débito como Crédito son 0"
                        )
                        
                except (ValueError, TypeError) as e:
                    resultado['errores'].append(f"Fila {idx + 2}: Error en montos - {e}")
                    resultado['valido'] = False
            
            # Calcular estadísticas
            df['Fecha'] = pd.to_datetime(df['Fecha'])
            df['Débito'] = pd.to_numeric(df['Débito'], errors='coerce').fillna(0)
            df['Crédito'] = pd.to_numeric(df['Crédito'], errors='coerce').fillna(0)
            
            resultado['estadisticas']['total_transacciones'] = len(df)
            resultado['estadisticas']['total_debitos'] = float(df['Débito'].sum())
            resultado['estadisticas']['total_creditos'] = float(df['Crédito'].sum())
            resultado['estadisticas']['fecha_inicio'] = df['Fecha'].min().strftime('%Y-%m-%d')
            resultado['estadisticas']['fecha_fin'] = df['Fecha'].max().strftime('%Y-%m-%d')
            resultado['estadisticas']['balance'] = (
                resultado['estadisticas']['total_creditos'] - 
                resultado['estadisticas']['total_debitos']
            )
            
            logger.info(f"Validación completada: {len(resultado['errores'])} errores, "
                       f"{len(resultado['advertencias'])} advertencias")
            
        except Exception as e:
            resultado['errores'].append(f"Error leyendo archivo: {str(e)}")
            resultado['valido'] = False
            logger.error(f"Error validando archivo PROGRAIN: {e}", exc_info=True)
        
        return resultado
