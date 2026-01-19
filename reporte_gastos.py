"""
Generador de Reportes de Gastos para EQUIPOS 4.0
Genera reportes en PDF y Excel con filtros avanzados: 
- Rango de fechas
- Equipo
- Cuenta contable
- Categoría
- Subcategoría
- Búsqueda de texto libre

Incluye totales, agrupaciones y formato profesional.
"""

import os
from datetime import datetime
from pathlib import Path
import logging

# PDF
from reportlab.lib import colors
from reportlab.lib. pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl. utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


class ReporteGastos:
    """Generador de reportes de gastos en PDF y Excel."""
    
    def __init__(self, datos_empresa:  dict | None = None, moneda_symbol: str = "RD$"):
        """
        Constructor. 
        
        Args:
            datos_empresa: Diccionario con información de la empresa
                {
                    "nombre":  str,
                    "direccion": str,
                    "telefono": str,
                    "email": str,
                    "logo_path": str (opcional)
                }
            moneda_symbol: Símbolo de la moneda (default: "RD$")
        """
        self.datos_empresa = datos_empresa or {}
        self.moneda_symbol = moneda_symbol
        self.styles = getSampleStyleSheet()
        
        # Estilos personalizados
        self.styles.add(ParagraphStyle(
            name='TituloReporte',
            parent=self.styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles. add(ParagraphStyle(
            name='SubtituloReporte',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors. HexColor('#666666'),
            spaceAfter=20,
            alignment=TA_CENTER
        ))
        
        self. styles.add(ParagraphStyle(
            name='InfoEmpresa',
            parent=self.styles['Normal'],
            fontSize=9,
            textColor=colors. HexColor('#333333'),
            alignment=TA_LEFT
        ))

    def generar_pdf(
        self,
        gastos:  list[dict],
        filtros_aplicados: dict,
        mapas:  dict,
        output_path: str,
        orientacion: str = "portrait"
    ) -> bool:
        """
        Genera un reporte PDF de gastos.
        
        Args:
            gastos: Lista de diccionarios con datos de gastos
            filtros_aplicados: Diccionario con los filtros usados
                {
                    "fecha_inicio": str,
                    "fecha_fin": str,
                    "equipo_nombre": str | None,
                    "cuenta_nombre": str | None,
                    "categoria_nombre": str | None,
                    "subcategoria_nombre": str | None,
                    "texto_busqueda": str | None
                }
            mapas: Diccionarios de mapeo
                {
                    "equipos": {id: nombre},
                    "cuentas": {id: nombre},
                    "categorias": {id: nombre},
                    "subcategorias": {id: nombre}
                }
            output_path:  Ruta del archivo PDF a generar
            orientacion: "portrait" o "landscape"
        
        Returns:
            True si se generó correctamente, False si hubo error
        """
        try:
            # Determinar tamaño de página
            pagesize = landscape(letter) if orientacion == "landscape" else letter
            
            # Crear documento
            doc = SimpleDocTemplate(
                output_path,
                pagesize=pagesize,
                rightMargin=0.5 * inch,
                leftMargin=0.5 * inch,
                topMargin=0.75 * inch,
                bottomMargin=0.5 * inch
            )
            
            story = []
            
            # Encabezado
            story.extend(self._crear_encabezado_pdf(filtros_aplicados))
            
            # Tabla de datos
            if gastos:
                story.append(self._crear_tabla_gastos_pdf(gastos, mapas, orientacion))
                story.append(Spacer(1, 0.2 * inch))
                
                # Totales
                story.extend(self._crear_totales_pdf(gastos))
            else:
                story.append(Paragraph(
                    "<para align='center'><i>No hay gastos que cumplan con los filtros aplicados.</i></para>",
                    self. styles['Normal']
                ))
            
            # Pie de página
            story.append(Spacer(1, 0.3 * inch))
            story.extend(self._crear_pie_pagina_pdf())
            
            # Construir PDF
            doc.build(story)
            logger.info(f"Reporte PDF generado:  {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error generando PDF de gastos: {e}", exc_info=True)
            return False

    def generar_excel(
        self,
        gastos: list[dict],
        filtros_aplicados: dict,
        mapas: dict,
        output_path: str
    ) -> bool:
        """
        Genera un reporte Excel de gastos.
        
        Args:
            gastos: Lista de gastos
            filtros_aplicados: Filtros aplicados
            mapas: Mapeos de IDs a nombres
            output_path:  Ruta del archivo Excel
        
        Returns:
            True si se generó correctamente
        """
        if not _HAS_OPENPYXL:
            logger. error("openpyxl no está instalado. No se puede generar Excel.")
            return False
        
        try: 
            wb = openpyxl. Workbook()
            ws = wb.active
            ws.title = "Gastos"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezado del reporte
            fila_actual = 1
            
            # Título
            ws.merge_cells(f'A{fila_actual}: I{fila_actual}')
            cell = ws[f'A{fila_actual}']
            cell.value = "REPORTE DE GASTOS"
            cell.font = Font(bold=True, size=16, color="1F4E78")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            fila_actual += 1
            
            # Empresa
            if self.datos_empresa. get("nombre"):
                ws.merge_cells(f'A{fila_actual}:I{fila_actual}')
                cell = ws[f'A{fila_actual}']
                cell.value = self.datos_empresa["nombre"]
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center')
                fila_actual += 1
            
            fila_actual += 1  # Espacio
            
            # Filtros aplicados
            ws[f'A{fila_actual}'] = "Filtros aplicados:"
            ws[f'A{fila_actual}'].font = Font(bold=True)
            fila_actual += 1
            
            ws[f'A{fila_actual}'] = "Período:"
            ws[f'B{fila_actual}'] = f"{filtros_aplicados.get('fecha_inicio', '')} al {filtros_aplicados.get('fecha_fin', '')}"
            fila_actual += 1
            
            if filtros_aplicados.get("equipo_nombre"):
                ws[f'A{fila_actual}'] = "Equipo:"
                ws[f'B{fila_actual}'] = filtros_aplicados["equipo_nombre"]
                fila_actual += 1
            
            if filtros_aplicados.get("cuenta_nombre"):
                ws[f'A{fila_actual}'] = "Cuenta:"
                ws[f'B{fila_actual}'] = filtros_aplicados["cuenta_nombre"]
                fila_actual += 1
            
            if filtros_aplicados.get("categoria_nombre"):
                ws[f'A{fila_actual}'] = "Categoría:"
                ws[f'B{fila_actual}'] = filtros_aplicados["categoria_nombre"]
                fila_actual += 1
            
            if filtros_aplicados.get("subcategoria_nombre"):
                ws[f'A{fila_actual}'] = "Subcategoría:"
                ws[f'B{fila_actual}'] = filtros_aplicados["subcategoria_nombre"]
                fila_actual += 1
            
            if filtros_aplicados.get("texto_busqueda"):
                ws[f'A{fila_actual}'] = "Búsqueda:"
                ws[f'B{fila_actual}'] = filtros_aplicados["texto_busqueda"]
                fila_actual += 1
            
            fila_actual += 1  # Espacio
            
            # Encabezados de tabla
            headers = ["Fecha", "Equipo", "Cuenta", "Categoría", "Subcategoría", "Descripción", "Monto", "Comentario", "Adjunto"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=fila_actual, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style
            
            fila_actual += 1
            fila_inicio_datos = fila_actual
            
            # Datos
            total_monto = 0.0
            equipos_mapa = mapas.get("equipos", {})
            cuentas_mapa = mapas.get("cuentas", {})
            categorias_mapa = mapas. get("categorias", {})
            subcategorias_mapa = mapas.get("subcategorias", {})
            
            for gasto in gastos:
                equipo_id = str(gasto.get("equipo_id", ""))
                cuenta_id = str(gasto. get("cuenta_id", ""))
                categoria_id = str(gasto.get("categoria_id", ""))
                subcategoria_id = str(gasto. get("subcategoria_id", ""))
                
                equipo_nombre = equipos_mapa.get(equipo_id, "Sin equipo") if equipo_id else "Sin equipo"
                cuenta_nombre = cuentas_mapa. get(cuenta_id, "")
                categoria_nombre = categorias_mapa.get(categoria_id, "")
                subcategoria_nombre = subcategorias_mapa.get(subcategoria_id, "")
                
                monto = float(gasto.get("monto", 0) or 0)
                total_monto += monto
                
                ws.cell(row=fila_actual, column=1).value = gasto.get("fecha", "")
                ws.cell(row=fila_actual, column=2).value = equipo_nombre
                ws.cell(row=fila_actual, column=3).value = cuenta_nombre
                ws.cell(row=fila_actual, column=4).value = categoria_nombre
                ws.cell(row=fila_actual, column=5).value = subcategoria_nombre
                ws.cell(row=fila_actual, column=6).value = gasto.get("descripcion", "")
                
                cell_monto = ws.cell(row=fila_actual, column=7)
                cell_monto. value = monto
                cell_monto. number_format = '#,##0.00'
                
                ws.cell(row=fila_actual, column=8).value = gasto.get("comentario", "")
                ws.cell(row=fila_actual, column=9).value = "Sí" if gasto.get("archivo_storage_path") else ""
                
                # Bordes
                for col in range(1, 10):
                    ws.cell(row=fila_actual, column=col).border = border_style
                
                fila_actual += 1
            
            # Fila de totales
            fila_actual += 1
            ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
            cell_total_label = ws[f'A{fila_actual}']
            cell_total_label.value = "TOTAL"
            cell_total_label.font = Font(bold=True, size=12)
            cell_total_label.alignment = Alignment(horizontal='right')
            
            cell_total_monto = ws.cell(row=fila_actual, column=7)
            cell_total_monto.value = total_monto
            cell_total_monto.number_format = f'"{self.moneda_symbol}" #,##0.00'
            cell_total_monto.font = Font(bold=True, size=12, color="1F4E78")
            cell_total_monto.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12  # Fecha
            ws.column_dimensions['B'].width = 20  # Equipo
            ws. column_dimensions['C'].width = 18  # Cuenta
            ws. column_dimensions['D'].width = 18  # Categoría
            ws.column_dimensions['E'].width = 20  # Subcategoría
            ws.column_dimensions['F']. width = 30  # Descripción
            ws.column_dimensions['G'].width = 15  # Monto
            ws.column_dimensions['H'].width = 30  # Comentario
            ws.column_dimensions['I'].width = 10  # Adjunto
            
            # Pie de página
            fila_actual += 3
            ws[f'A{fila_actual}'] = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{fila_actual}'].font = Font(italic=True, size=9, color="666666")
            
            # Guardar
            wb.save(output_path)
            logger.info(f"Reporte Excel generado:  {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error generando Excel de gastos: {e}", exc_info=True)
            return False

    def _crear_encabezado_pdf(self, filtros:  dict) -> list:
        """Crea el encabezado del PDF con información de empresa y filtros."""
        elements = []
        
        # Logo (si existe)
        logo_path = self.datos_empresa.get("logo_path")
        if logo_path and os.path.exists(logo_path):
            try:
                img = Image(logo_path, width=1.5*inch, height=0.75*inch)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 0.1*inch))
            except Exception as e:
                logger.warning(f"No se pudo cargar logo: {e}")
        
        # Nombre empresa
        if self.datos_empresa.get("nombre"):
            elements.append(Paragraph(self.datos_empresa["nombre"], self.styles['InfoEmpresa']))
        
        if self.datos_empresa.get("direccion"):
            elements.append(Paragraph(self.datos_empresa["direccion"], self.styles['InfoEmpresa']))
        
        if self.datos_empresa.get("telefono") or self.datos_empresa.get("email"):
            contacto = " | ".join(filter(None, [
                self.datos_empresa.get("telefono"),
                self. datos_empresa.get("email")
            ]))
            elements. append(Paragraph(contacto, self.styles['InfoEmpresa']))
        
        elements.append(Spacer(1, 0.2*inch))
        
        # Título
        elements.append(Paragraph("REPORTE DE GASTOS", self.styles['TituloReporte']))
        
        # Período
        periodo = f"Período: {filtros.get('fecha_inicio', '')} al {filtros.get('fecha_fin', '')}"
        elements. append(Paragraph(periodo, self.styles['SubtituloReporte']))
        
        # Filtros adicionales
        filtros_texto = []
        if filtros.get("equipo_nombre"):
            filtros_texto.append(f"<b>Equipo:</b> {filtros['equipo_nombre']}")
        if filtros. get("cuenta_nombre"):
            filtros_texto.append(f"<b>Cuenta:</b> {filtros['cuenta_nombre']}")
        if filtros. get("categoria_nombre"):
            filtros_texto.append(f"<b>Categoría: </b> {filtros['categoria_nombre']}")
        if filtros.get("subcategoria_nombre"):
            filtros_texto. append(f"<b>Subcategoría:</b> {filtros['subcategoria_nombre']}")
        if filtros.get("texto_busqueda"):
            filtros_texto.append(f"<b>Búsqueda:</b> {filtros['texto_busqueda']}")
        
        if filtros_texto:
            elements.append(Paragraph(" | ".join(filtros_texto), self.styles['Normal']))
            elements.append(Spacer(1, 0.15*inch))
        
        return elements

    def _crear_tabla_gastos_pdf(self, gastos: list[dict], mapas: dict, orientacion: str) -> Table:
        """Crea la tabla de gastos para PDF."""
        equipos_mapa = mapas. get("equipos", {})
        cuentas_mapa = mapas.get("cuentas", {})
        categorias_mapa = mapas.get("categorias", {})
        subcategorias_mapa = mapas.get("subcategorias", {})
        
        # Encabezados
        if orientacion == "landscape":
            headers = ["Fecha", "Equipo", "Cuenta", "Categoría", "Subcategoría", "Descripción", "Monto", "Comentario"]
            col_widths = [0.8*inch, 1.2*inch, 1*inch, 1*inch, 1.2*inch, 2*inch, 0.9*inch, 1.5*inch]
        else:
            headers = ["Fecha", "Equipo", "Cuenta", "Categoría", "Monto", "Descripción"]
            col_widths = [0.8*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1*inch, 2*inch]
        
        data = [headers]
        
        # Filas de datos
        for gasto in gastos:
            equipo_id = str(gasto. get("equipo_id", ""))
            cuenta_id = str(gasto.get("cuenta_id", ""))
            categoria_id = str(gasto.get("categoria_id", ""))
            subcategoria_id = str(gasto.get("subcategoria_id", ""))
            
            equipo_nombre = equipos_mapa.get(equipo_id, "Sin equipo") if equipo_id else "Sin equipo"
            cuenta_nombre = cuentas_mapa.get(cuenta_id, "")
            categoria_nombre = categorias_mapa.get(categoria_id, "")
            subcategoria_nombre = subcategorias_mapa.get(subcategoria_id, "")
            
            monto = float(gasto.get("monto", 0) or 0)
            monto_str = f"{self.moneda_symbol} {monto: ,.2f}"
            
            if orientacion == "landscape":
                fila = [
                    gasto.get("fecha", ""),
                    equipo_nombre[: 20],
                    cuenta_nombre[:15],
                    categoria_nombre[: 15],
                    subcategoria_nombre[:20],
                    gasto.get("descripcion", "")[:30],
                    monto_str,
                    gasto.get("comentario", "")[:25]
                ]
            else: 
                fila = [
                    gasto.get("fecha", ""),
                    equipo_nombre[:25],
                    cuenta_nombre[: 20],
                    categoria_nombre[: 20],
                    monto_str,
                    gasto.get("descripcion", "")[:35]
                ]
            
            data.append(fila)
        
        # Crear tabla
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Estilo
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Fecha centrada
            ('ALIGN', (-2, 1), (-2, -1), 'RIGHT'),  # Monto a la derecha
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            
            # Filas alternas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        
        return table

    def _crear_totales_pdf(self, gastos:  list[dict]) -> list:
        """Crea la sección de totales."""
        elements = []
        
        total_monto = sum(float(g.get("monto", 0) or 0) for g in gastos)
        
        # Tabla de totales
        data = [
            ["Total de Gastos:", str(len(gastos))],
            ["Monto Total:", f"{self.moneda_symbol} {total_monto: ,.2f}"]
        ]
        
        table = Table(data, colWidths=[2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#1F4E78')),
            ('LINEABOVE', (0, 1), (-1, 1), 1, colors.black),
        ]))
        
        table.hAlign = 'RIGHT'
        elements.append(table)
        
        return elements

    def _crear_pie_pagina_pdf(self) -> list:
        """Crea el pie de página."""
        elements = []
        
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        texto = f"<para align='center'><font size='8' color='#666666'>Generado el {fecha_hora} | EQUIPOS 4.0</font></para>"
        
        elements.append(Paragraph(texto, self.styles['Normal']))
        
        return elements